#! /usr/bin/python

#import pandas as pd
import openpyxl as ox
import os
import re

def main():
# Open Excel file and read that file.
    file = r'C:\Users\LENOVO\Desktop\V7 Answer key.xlsx'
    book = ox.load_workbook(file)
    sheet = book.active
    
   # b2 = sheet['B2']
  #  a2 = sheet['A2']
 #   a3 = sheet.cell(row=3, column=2)
    cells = sheet['A2': 'C41']
    c=sheet['B2': 'B41']
    l=[]
    ans=[]
    demo=[]
    split_ans=[]
    
#    print(b2.value)
 #   print(a2.value)
 #   print(a3.value)
   
        
    for c1, c2,c3 in cells:
     #   a=("{0:8} {1:8} {2:8}".format(c1.value, c2.value,c3.value))
     #   print(a)
        l.append(c2.value)
        
#    print(l)
    answerkey = 0
    for k in l:
        print(k)
        answerkey += 1
    print "Total count of answerkey :", answerkey 
        
        
        
    #response={}
 
 #open text file and read the content
    f = open("C:\Users\LENOVO\Desktop\Va.txt","r")
 #   print(f.readlines())
    for i in f:
#        print(i)
        ans.append(i)
        demo.append(i.split(','))
        
        
 #   print(ans)
  #  print(demo)
   # print(type(l))
    print(demo[0][1])
    correct = 0
    uncorrect = 0
    unanswered = 0
#    element = 0
    
    for elem in demo[0][1]:
#        element += 1
#    print "Total num of elem : ", element
        print(elem)
        
        if(elem == k) :
  #          print("Answer is correct")
            #    correct.append(elem)
            correct += 1
           
            
             
        elif(elem != k):
 #               print("incorrect ans")
                #    unanswered.append(elem)
                uncorrect += 1
                 
        else:
 #               print("unanswered")
                unanswered += 1
                  #  count+=1
                   # unanswered.append('Unanswered -',+str(count))
    
    print "Total correct answer are :", correct
    print "Total incorrect answer are : ", uncorrect
    print "Total unanswered answer are : ", unanswered
            
    
    # code to split it into 2 lists 
   # res1, res2 = map(list, zip(*demo)) 
  
    # printing result 
    #print("final lists", res1, "\n", res2) 
    
    #for i in res2:
     #   split_ans.append(i.split())
    
    
    #print(split_ans)
    
   
    
    
    
    
 #   nn=3
  #  xf = [ans[ir:ir + nn] for ir in range(0, len(ans), nn)]  
   # print(xf) 
     #J=ans.splitsplit(", ")

      #print(J)
   
    
#    print(file)
#    xl=pd.ExcelFile(file)
 #   print(xl.sheet_names)
  #  a = xl.sheet_names()
   # print(a)
#    print(xl.sheet_names)
#    df1 = xl.parse('Sheet1')

#    a=df1.sheet['A1'].value
 #   print(a)
 #   for i in range(1,4):
 #       print(i,sheet.cell(row=i,column=3).values)
    
#	print("Hello")


if __name__=='__main__':
	main()
    